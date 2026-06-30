#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
爬取科技部政务服务平台 (fuwu.most.gov.cn) 的人类遗传资源审批结果历史数据
覆盖范围：2021年 ~ 2023年第18批（早于现有HGRS API数据）

用法：
    python hgr_most_scraper.py

配置：从 ../config.ini 读取 data_dir（与 hgr_main.py 共用同一配置）
"""

import sys, io, os, re, time, logging, shutil, configparser
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import requests
    from lxml import html as lxml_html
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook, Workbook
except ImportError as e:
    print(f"❌ 缺少依赖: {e}")
    print("安装: pip install requests lxml openpyxl")
    sys.exit(1)

# 复用 excel_writer 的格式定义
sys.path.insert(0, os.path.dirname(__file__))
from excel_writer import CATEGORY_SHEETS, COL_WIDTHS

# 从 config.ini 读取输出目录配置
_config = configparser.ConfigParser()
_config.read(os.path.join(os.path.dirname(__file__), '..', 'config.ini'), encoding='utf-8')
_data_dir_rel = _config.get('DEFAULT', 'data_dir', fallback='./data')
_summary_filename = _config.get('DEFAULT', 'summary_filename', fallback='汇总_中国人类遗传资源行政许可事项.xlsx')
_config_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DATA_DIR = os.path.abspath(os.path.join(_config_dir, _data_dir_rel))
SUMMARY_PATH = os.path.join(DATA_DIR, _summary_filename)
BATCHES_DIR = os.path.join(DATA_DIR, 'batches')
os.makedirs(DATA_DIR, exist_ok=True)

BASE_URL = 'https://fuwu.most.gov.cn'

MAX_PAGES = 25
REQUEST_DELAY = 0.5
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

# ============ 日志 ============
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger('most')

# ============ Excel 样式（独立于 excel_writer 的 HGRWriter 实例属性） ============
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='微软雅黑', size=9)
DATA_ALIGN = Alignment(vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

# ============ 中文数字映射 ============
CN_NUMS = {'零':0,'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
           '十一':11,'十二':12,'十三':13,'十四':14,'十五':15,'十六':16,'十七':17,'十八':18,'十九':19,
           '二十':20,'二十一':21,'二十二':22,'二十三':23,'二十四':24,'二十五':25,'廿':20,'三十':30}

def parse_cn_batch(text):
    """从文本中提取年份和批次（支持中文数字）"""
    year_m = re.search(r'(\d{4})年', text)
    year = int(year_m.group(1)) if year_m else 0
    # 尝试阿拉伯数字
    batch_m = re.search(r'第(\d+)批', text)
    if batch_m:
        return year, int(batch_m.group(1))
    # 尝试中文数字
    for cn, num in sorted(CN_NUMS.items(), key=lambda x: -len(x[0])):
        if re.search(rf'第{cn}批', text):
            return year, num
    return year, 0

# ============ 请求 ============
session = requests.Session()
session.headers.update({'User-Agent': USER_AGENT})

def fetch(url):
    try:
        r = session.get(url, timeout=30)
        r.encoding = 'utf-8'
        if r.status_code == 200:
            return r.text
        log.warning(f"HTTP {r.status_code}: {url}")
        return None
    except Exception as e:
        log.error(f"请求失败 {url}: {e}")
        return None

# ============ 解析 ============
def parse_list_page(html_text):
    """提取列表页中的审批结果链接"""
    doc = lxml_html.fromstring(html_text)
    links = doc.xpath('//a[contains(@href, "/html/tztg/xzxkzx/")]')
    items = []
    for a in links:
        text = (a.text_content() or '').strip().replace('\xa0', ' ')
        href = a.get('href', '')
        if ('审批结果' in text and 
            not any(kw in text for kw in ['简化流程', '审查结果', '备案情况', '备案公示'])):
            full_url = href if href.startswith('http') else BASE_URL + href
            year, batch_num = parse_cn_batch(text)
            items.append({'title': text, 'url': full_url, 'year': year, 'batch': batch_num})
    return items

def classify_approval(no):
    if 'CJ' in no: return '采集审批'
    if 'BC' in no: return '保藏审批'
    if 'GH' in no: return '国际科学研究合作审批'
    if 'CC' in no: return '材料出境证明'
    return '国际科学研究合作审批'

def parse_detail_page(html_text, title):
    """提取详情页表格，返回 {类别: [{列名:值}]}"""
    doc = lxml_html.fromstring(html_text)
    tables = doc.xpath('//table')
    if not tables:
        return None
    result = {cn: [] for cn, _ in CATEGORY_SHEETS}
    for table in tables:
        rows = table.xpath('.//tr')
        if len(rows) < 2:
            continue
        header_cols = len(rows[0].xpath('.//td|.//th'))
        data_rows = []
        for row in rows[1:]:
            cells = [c.text_content().strip().replace('\xa0', ' ').replace('\n', ' ').replace('\r', ' ')
                     for c in row.xpath('.//td|.//th')]
            cells = [re.sub(r'  +', ' ', c).strip() for c in cells]
            if not any(cells) or cells[0] in ('', '序号') or '序号' in cells[0]:
                continue
            data_rows.append(cells)
        if not data_rows:
            continue
        if header_cols <= 5:
            for cells in data_rows:
                approval_no = cells[1] if len(cells) > 1 else ''
                cat = classify_approval(approval_no)
                result[cat].append({
                    '序号': cells[0] if len(cells) > 0 else '',
                    '批次': title,
                    '审批号': approval_no,
                    '项目名称': cells[2] if len(cells) > 2 else '',
                    '申请单位': cells[3] if len(cells) > 3 else '',
                    '批准时间': cells[4] if len(cells) > 4 else '',
                })
        elif header_cols >= 8:
            for cells in data_rows:
                approval_no = cells[1] if len(cells) > 1 else ''
                result['国际科学研究合作审批'].append({
                    '序号': cells[0] if len(cells) > 0 else '',
                    '批次': title,
                    '审批号': approval_no,
                    '项目名称': cells[2] if len(cells) > 2 else '',
                    '医疗机构(组长单位)': cells[3] if len(cells) > 3 else '',
                    '申办方': cells[4] if len(cells) > 4 else '',
                    '合同研究组织': cells[5] if len(cells) > 5 else '',
                    '检测/数据单位': cells[6] if len(cells) > 6 else '',
                    '批准时间': cells[7] if len(cells) > 7 else '',
                })
    return result

def batch_sort_key(batch_str):
    y, b = parse_cn_batch(str(batch_str))
    return (y, b)

# ============ Excel 写入 ============
def style_cell(cell, header=False, row_idx=None, center=False):
    cell.border = THIN_BORDER
    if header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    else:
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN if center else DATA_ALIGN
        if row_idx and row_idx % 2 == 0:
            cell.fill = ALT_FILL

def create_batch_excel(batch_data, output_dir):
    """创建单批次独立Excel文件"""
    year, batch_num, title = batch_data['year'], batch_data['batch'], batch_data['title']
    filename = f"中国人类遗传资源行政许可事项{year}年第{batch_num}批审批结果.xlsx"
    path = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for cat_name, headers in CATEGORY_SHEETS:
        rows = batch_data['data'].get(cat_name, [])
        if not rows:
            continue
        ws = wb.create_sheet(title=cat_name)
        for ci, hn in enumerate(headers, 1):
            c = ws.cell(1, ci, hn)
            style_cell(c, header=True)
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(hn, 20)
        for ri, row in enumerate(rows, 2):
            for ci, hn in enumerate(headers, 1):
                c = ws.cell(ri, ci, row.get(hn, ''))
                style_cell(c, row_idx=ri - 1, center=hn in ['序号', '批准时间', '批次'])
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows) + 1}'
    wb.save(path)
    log.info(f"  ✅ 独立文件: {filename}")
    return path

def merge_into_summary(all_batch_data):
    """将MOST批次合并到现有汇总Excel"""
    # 备份
    if os.path.exists(SUMMARY_PATH):
        bak = SUMMARY_PATH.replace('.xlsx', f'.most_bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        shutil.copy2(SUMMARY_PATH, bak)
        log.info(f"📦 备份: {os.path.basename(bak)}")
    
    wb = load_workbook(SUMMARY_PATH) if os.path.exists(SUMMARY_PATH) else Workbook()
    
    for cat_name, headers in CATEGORY_SHEETS:
        if cat_name not in wb.sheetnames:
            ws = wb.create_sheet(title=cat_name)
            for ci, hn in enumerate(headers, 1):
                c = ws.cell(1, ci, hn)
                style_cell(c, header=True)
                ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(hn, 20)
            ws.freeze_panes = 'A2'
        ws = wb[cat_name]
        
        # 读取现有数据
        existing = []
        for r in range(2, ws.max_row + 1):
            row = {}
            has_data = False
            for ci, hn in enumerate(headers, 1):
                val = ws.cell(r, ci).value
                if val is not None:
                    has_data = True
                row[hn] = val if val is not None else ''
            if has_data:
                existing.append(row)
        
        # 新增数据
        new_rows = []
        for bd in all_batch_data:
            new_rows.extend(bd['data'].get(cat_name, []))
        
        # 合并 & 去重
        all_rows = existing + new_rows
        seen_approval = set()
        deduped = []
        for row in all_rows:
            key = row.get('审批号', '')
            if key and key not in seen_approval:
                seen_approval.add(key)
                deduped.append(row)
            elif not key:
                deduped.append(row)
        
        # 排序
        deduped.sort(key=lambda r: (batch_sort_key(r.get('批次', '')), str(r.get('序号', '999999')).zfill(6)))
        
        # 重新编号
        for i, row in enumerate(deduped, 1):
            row['序号'] = i
        
        # 清空并重写
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).value = None
        for ri, row in enumerate(deduped, 2):
            for ci, hn in enumerate(headers, 1):
                c = ws.cell(ri, ci, row.get(hn, ''))
                style_cell(c, row_idx=ri - 1, center=hn in ['序号', '批准时间', '批次'])
        
        log.info(f"  📊 {cat_name}: {len(existing)} → {len(deduped)} 条 ({len(new_rows)} 新增)")
        if deduped:
            ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(deduped) + 1}'
    
    wb.save(SUMMARY_PATH)
    log.info(f"✅ 汇总已更新: {SUMMARY_PATH}")

# ============ 主流程 ============
def main():
    log.info("=" * 60)
    log.info(f"🚀 科技部政务服务平台历史审批结果爬取")
    log.info(f"数据目录: {DATA_DIR}")
    
    # 阶段1：遍历列表页
    log.info("\n【阶段1】遍历列表页")
    all_items = []
    for pg in range(1, MAX_PAGES + 1):
        url = BASE_URL + '/html/jgcx/index' + ('' if pg == 1 else f'_{pg}') + '.html'
        html = fetch(url)
        if not html:
            break
        items = parse_list_page(html)
        log.info(f"  第{pg}页: {len(items)} 个审批结果")
        all_items.extend(items)
        doc = lxml_html.fromstring(html)
        list_links = doc.xpath('//a[contains(@href, "/html/tztg/xzxkzx/")]')
        if len(list_links) == 0:
            log.info(f"  ℹ️ 第{pg}页无列表项，停止")
            break
        time.sleep(REQUEST_DELAY)
    
    # 去重
    seen_urls = set()
    unique_items = []
    for item in all_items:
        if item['url'] not in seen_urls:
            seen_urls.add(item['url'])
            unique_items.append(item)
    
    unique_items.sort(key=lambda x: (x['year'], x['batch']))
    log.info(f"\n📋 共 {len(unique_items)} 个批次 ({unique_items[0]['title']} ~ {unique_items[-1]['title']})")
    
    # 阶段2：提取详情页
    log.info("\n【阶段2】提取详情页")
    valid_batches, errors = [], []
    for idx, item in enumerate(unique_items):
        log.info(f"  [{idx+1}/{len(unique_items)}] {item['title']}")
        html = fetch(item['url'])
        if not html:
            errors.append(item['title'])
            continue
        data = parse_detail_page(html, item['title'])
        if not data or sum(len(v) for v in data.values()) == 0:
            errors.append(item['title'])
            continue
        total = sum(len(v) for v in data.values())
        log.info(f"    ✅ {total} 条: {{k:len(v) for k,v in data.items() if v}}")
        valid_batches.append({'year': item['year'], 'batch': item['batch'],
                               'title': item['title'], 'data': data, 'total_count': total})
        time.sleep(REQUEST_DELAY)
    
    log.info(f"\n✅ {len(valid_batches)} 批次成功, ❌ {len(errors)} 失败")
    by_cat = {}
    for b in valid_batches:
        for cat, rows in b['data'].items():
            by_cat[cat] = by_cat.get(cat, 0) + len(rows)
    log.info(f"📊 总记录: {sum(b['total_count'] for b in valid_batches)}")
    for cat, cnt in by_cat.items():
        log.info(f"  {cat}: {cnt}")
    
    # 阶段3：写入Excel
    log.info("\n【阶段3】写入Excel")
    for b in valid_batches:
        create_batch_excel(b, BATCHES_DIR)
    merge_into_summary(valid_batches)
    
    log.info(f"\n{'='*60}")
    log.info("🏁 完成！")

if __name__ == '__main__':
    main()
