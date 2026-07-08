"""汇总 Excel 重新排序 + 编号"""
import openpyxl
import re
import sys
import io
import os
import configparser

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

_config = configparser.ConfigParser()
_config.read(os.path.join(os.path.dirname(__file__), '..', 'config.ini'), encoding='utf-8')
_data_dir_rel = _config.get('DEFAULT', 'data_dir', fallback='./data')
_summary_filename = _config.get('DEFAULT', 'summary_filename', fallback='汇总_中国人类遗传资源行政许可事项.xlsx')
_config_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
summary_path = os.path.join(os.path.abspath(os.path.join(_config_dir, _data_dir_rel)), _summary_filename)

wb = openpyxl.load_workbook(summary_path)

sheets_config = {
    '采集审批': {'sort_col': 5},
    '保藏审批': {'sort_col': 5},
    '国际科学研究合作审批': {'sort_col': 8},
    '材料出境证明': {'sort_col': 5},
}

def batch_sort_key(r):
    batch_str = str(r[1]) if r[1] else ''
    m = re.search(r'(\d{4})年.*?(\d+)批', batch_str)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return (0, 0)

for sheet_name in wb.sheetnames:
    if sheet_name not in sheets_config:
        continue

    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows.append(row)

    if not rows:
        print(f'{sheet_name}: 空')
        continue

    # 按批次排序
    sorted_rows = sorted(rows, key=batch_sort_key)

    # 重新编号
    for i, row in enumerate(sorted_rows):
        row_list = list(row)
        row_list[0] = i + 1
        sorted_rows[i] = tuple(row_list)

    # 写回
    for i, row in enumerate(sorted_rows):
        for j, val in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=val)

    print(f'{sheet_name}: {len(rows)} 条 -> 排序 + 重新编号 ✅')

wb.save(summary_path)
print(f'\n✅ 汇总文件排序完成: {summary_path}')
