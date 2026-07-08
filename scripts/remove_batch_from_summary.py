"""从汇总 Excel 中移除指定批次的数据"""
import openpyxl
import sys
import io

summary_path = r'C:\Users\Owen\.copaw\workspaces\default\skills\HGRS\data\汇总_中国人类遗传资源行政许可事项.xlsx'
target_batch = '2026年第12批'

wb = openpyxl.load_workbook(summary_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # 读取所有行，跳过表头
    rows_to_keep = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # 批次列是第2列（index 1）
        batch_str = str(row[1]) if row[1] else ''
        if target_batch in batch_str:
            continue
        rows_to_keep.append(row)

    # 清空旧数据
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 写回保留的数据，重新编号
    for i, row in enumerate(rows_to_keep):
        row_list = list(row)
        row_list[0] = i + 1  # 重新编号
        for j, val in enumerate(row_list):
            ws.cell(row=i+2, column=j+1, value=val)

    removed = len(list(ws.iter_rows(min_row=2, values_only=True))) - len(rows_to_keep)
    print(f'{sheet_name}: 保留 {len(rows_to_keep)} 条, 移除 {removed} 条 ✅')

wb.save(summary_path)
print(f'\n✅ 已从汇总文件中移除 {target_batch} 数据')
