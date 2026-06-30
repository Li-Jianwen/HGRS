#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HGR Excel Writer
输出Excel：汇总文件（方案C：每个审批类别一个Sheet，包含所有批次）+ 独立批次文件
"""

import os
import sys
import io
import logging
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 表头定义（按方案C顺序）
CATEGORY_SHEETS = [
    ('采集审批', ['序号', '批次', '审批号', '项目名称', '申请单位', '批准时间']),
    ('保藏审批', ['序号', '批次', '审批号', '项目名称', '申请单位', '批准时间']),
    ('国际科学研究合作审批', ['序号', '批次', '审批号', '项目名称', '医疗机构(组长单位)', '申办方', '合同研究组织', '检测/数据单位', '批准时间']),
    ('材料出境证明', ['序号', '批次', '审批号', '项目名称', '申请单位', '批准时间']),
]

# 列宽定义
COL_WIDTHS = {
    '序号': 8,
    '批次': 16,
    '审批号': 22,
    '项目名称': 50,
    '申请单位': 30,
    '医疗机构(组长单位)': 28,
    '申办方': 32,
    '合同研究组织': 30,
    '检测/数据单位': 32,
    '批准时间': 15,
}

class HGRWriter:
    def __init__(self, data_dir: str, summary_filename: str, batch_template: str, logger=None):
        self.data_dir = data_dir
        self.summary_path = os.path.join(data_dir, summary_filename)
        self.batch_template = batch_template
        self.logger = logger or logging.getLogger(__name__)
        
        # Styles
        self.header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.data_font = Font(name='微软雅黑', size=9)
        self.data_alignment = Alignment(vertical='center', wrap_text=True)
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    def _apply_cell_style(self, cell, is_header: bool, row_idx: int, is_center: bool = False):
        """
        应用单元格样式
        """
        if is_header:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        else:
            cell.font = self.data_font
            cell.border = self.thin_border
            if is_center:
                cell.alignment = self.center_alignment
            else:
                cell.alignment = self.data_alignment
            # 偶数行交替颜色
            if row_idx % 2 == 0:
                cell.fill = self.alt_fill

    def create_batch_file(self, batch_data: Dict, output_dir: str) -> str:
        """
        创建单批次独立文件（保留原始结构）
        返回: 输出文件路径
        """
        year = batch_data['year']
        batch = batch_data['batch']
        title = batch_data['title']
        data = batch_data['data']  # {category: [rows...]}
        
        filename = self.batch_template.format(year=year, batch=batch)
        output_path = os.path.join(output_dir, filename)
        os.makedirs(output_dir, exist_ok=True)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create sheets for each category that has data
        for cat_name, _ in CATEGORY_SHEETS:
            if not data.get(cat_name) or len(data[cat_name]) == 0:
                continue
            
            ws = wb.create_sheet(title=cat_name)
            headers = [h for (c, h) in CATEGORY_SHEETS if c == cat_name][0]
            ncols = len(headers)
            
            # Write header
            for col_idx, header_name in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header_name)
                self._apply_cell_style(cell, is_header=True, row_idx=1)
            
            # Write data
            for row_idx, row_dict in enumerate(data[cat_name], 2):
                for col_idx, header_name in enumerate(headers, 1):
                    value = row_dict.get(header_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    is_center = header_name in ['序号', '批准时间', '批次']
                    self._apply_cell_style(cell, is_header=False, row_idx=row_idx, is_center=is_center)
            
            # Set column widths
            for col_idx, header_name in enumerate(headers, 1):
                width = COL_WIDTHS.get(header_name, 20)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Freeze header
            ws.freeze_panes = 'A2'
            # Auto filter
            ws.auto_filter.ref = f'A1:{get_column_letter(ncols)}{len(data[cat_name]) + 1}'
        
        wb.save(output_path)
        self.logger.info(f"✅ 独立批次文件保存: {output_path}")
        return output_path

    def get_current_summary(self) -> Optional[Workbook]:
        """
        读取现有的汇总Excel，如果不存在返回None
        """
        if not os.path.exists(self.summary_path):
            return None
        try:
            wb = load_workbook(self.summary_path)
            self.logger.info(f"📂 读取现有汇总文件: {self.summary_path}")
            return wb
        except Exception as e:
            self.logger.warning(f"⚠️ 读取现有汇总文件失败，将创建新文件: {str(e)}")
            return None

    def create_new_summary(self) -> Workbook:
        """
        创建新的汇总Excel
        """
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create sheets for each category
        for cat_name, headers in CATEGORY_SHEETS:
            ws = wb.create_sheet(title=cat_name)
            # Write header
            for col_idx, header_name in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header_name)
                self._apply_cell_style(cell, is_header=True, row_idx=1)
            # Set column widths
            for col_idx, header_name in enumerate(headers, 1):
                width = COL_WIDTHS.get(header_name, 20)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.freeze_panes = 'A2'
        
        self.logger.info("📝 创建新汇总文件")
        return wb

    def insert_new_batch(self, wb: Workbook, batch_data: Dict) -> None:
        """
        将新批次插入到每个分类Sheet的表头下方（最新批次在最前面）
        """
        data = batch_data['data']
        batch_year = batch_data['year']
        batch_batch = batch_data['batch']
        batch_str = f"{batch_year}年第{batch_batch}批"
        
        for cat_name, headers in CATEGORY_SHEETS:
            if cat_name not in wb.sheetnames:
                continue
            
            ws = wb[cat_name]
            cat_data = data.get(cat_name, [])
            if not cat_data:
                continue
            
            nrows = ws.max_row
            if nrows >= 1:
                # Insert rows after header (row 2 is inserted after header)
                ws.insert_rows(2, len(cat_data))
            else:
                # No existing data, start from row 2
                pass
            
            for row_offset, row_dict in enumerate(cat_data):
                excel_row = 2 + row_offset
                for col_idx, header_name in enumerate(headers, 1):
                    value = row_dict.get(header_name, '')
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    is_center = header_name in ['序号', '批准时间', '批次']
                    self._apply_cell_style(cell, is_header=False, row_idx=excel_row, is_center=is_center)
        
        self.logger.info(f"✅ 新批次 {batch_str} 已插入汇总文件")

    def write_summary(self, wb: Workbook) -> None:
        """
        保存汇总文件
        """
        # Update auto-filter on each sheet
        for cat_name, headers in CATEGORY_SHEETS:
            if cat_name not in wb.sheetnames:
                continue
            ws = wb[cat_name]
            nrows = ws.max_row
            ncols = len(headers)
            if nrows >= 1:
                ws.auto_filter.ref = f'A1:{get_column_letter(ncols)}{nrows}'
        
        wb.save(self.summary_path)
        self.logger.info(f"✅ 汇总文件保存: {self.summary_path}")

    def process_new_batch(self, batch_data: Dict, batches_dir: str) -> Dict:
        """
        完整处理新批次：
        1. 创建独立文件
        2. 读取汇总文件（或新建）
        3. 插入新批次到汇总
        4. 保存汇总
        返回路径信息
        """
        # 1. 独立文件
        batch_path = self.create_batch_file(batch_data, batches_dir)
        
        # 2. 汇总文件
        wb = self.get_current_summary()
        if wb is None:
            wb = self.create_new_summary()
        
        # 3. 插入新批次
        self.insert_new_batch(wb, batch_data)
        
        # 4. 保存
        self.write_summary(wb)
        
        return {
            'batch_path': batch_path,
            'summary_path': self.summary_path,
            'batch_count': {cat: len(batch_data['data'][cat]) for cat in batch_data['data'] if batch_data['data'][cat]},
            'total_count': batch_data['total_count'],
        }


if __name__ == '__main__':
    # 测试：读取一个已处理的批次数据，写入Excel
    import argparse
    import json
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--batch-json', required=True, help='处理后的批次数据JSON')
    parser.add_argument('--data-dir', default='../data', help='数据目录')
    args = parser.parse_args()
    
    with open(args.batch_json, 'r', encoding='utf-8') as f:
        batch_data = json.load(f)
    
    writer = HGRWriter(
        data_dir=args.data_dir,
        summary_filename='汇总_中国人类遗传资源行政许可事项.xlsx',
        batch_template='中国人类遗传资源行政许可事项{year}年第{batch}批审批结果公示.xlsx'
    )
    
    batches_dir = os.path.join(args.data_dir, 'batches')
    result = writer.process_new_batch(batch_data, batches_dir)
    print(f"\n✅ 处理完成")
    print(f"独立文件: {result['batch_path']}")
    print(f"汇总文件: {result['summary_path']}")
    print(f"记录统计: {result['batch_count']}")
    print(f"总记录: {result['total_count']}")
